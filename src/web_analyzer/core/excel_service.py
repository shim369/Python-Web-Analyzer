import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from web_analyzer.models import ScrapingJob, SiteAssessment


class ExcelService:
    """Excelファイルのパースおよび生成を担当するサービス。"""

    @staticmethod
    def import_excel(
        file_path: Path,
        operator_name: str,
        threshold_1: int = 10,
        threshold_2: int = 15,
        threshold_3: int = 20,
    ) -> tuple[ScrapingJob, list[SiteAssessment]]:
        """Excelからドメイン名(B列)を抽出し、パラメータを紐付けてモデルを生成する。"""
        job_id = str(uuid.uuid4())
        job = ScrapingJob(
            id=job_id,
            operator_name=operator_name,
            threshold_1=threshold_1,
            threshold_2=threshold_2,
            threshold_3=threshold_3,
            status="processing",
        )

        assessments: list[SiteAssessment] = []
        wb = load_workbook(str(file_path), data_only=True)
        ws = wb.active

        assert isinstance(ws, Worksheet)

        # アプリ実行時の日付 (例: "7/14") を取得
        current_date_str = datetime.now().strftime("%m/%d").lstrip("0").replace("/0", "/")

        # 2行目から開始（1行目はヘッダーを想定。インポート元のB列にサイト名がある前提）
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=2).value

            if cell_value is None:
                cell_value = ws.cell(row=row, column=1).value

            if cell_value is None:
                continue

            domain = str(cell_value).strip()
            if not domain or domain == "サイト名":
                continue

            assessment = SiteAssessment(
                id=str(uuid.uuid4()),
                job_id=job_id,
                date_str=current_date_str,
                domain_name=domain,
                operator_name=operator_name,
            )
            assessments.append(assessment)

        return job, assessments

    @staticmethod
    def export_excel(assessments: list[SiteAssessment], output_path: Path) -> None:
        """指定された解析結果リストをスタイリッシュなデザインでExcel出力する。"""
        wb = Workbook()
        ws = wb.active
        assert isinstance(ws, Worksheet)
        ws.title = "調査結果"

        # グリッド線（目盛線）を常に表示する設定
        ws.views.sheetView[0].showGridLines = True

        # スタイル定義
        font_family = "Yu Gothic"

        # フォント設定
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)

        # 背景色設定（シックなネイビー）
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # 罫線設定（極細のライトグレー）
        thin_border_side = Side(style="thin", color="D9D9D9")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # 新仕様に基づいたヘッダー（A〜O列）
        headers = [
            "日付",
            "サイト名",
            "調査結果",
            "SSLあり",
            "SSL常時",
            "階層数",
            "svcmd",
            "構成",
            "ページ数",
            "使用CMS",
            "用途",
            "問合せ項目",
            "不可の理由",
            "備考",
            "担当",
        ]
        ws.append(headers)

        # ヘッダー行のデザイン適用
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # ヘッダーの行高を少し広げてゆとりを持たせる
        ws.row_dimensions[1].height = 26

        # データのマッピング書き込み
        for item in assessments:
            row_data = [
                item.date_str,
                item.domain_name,
                item.evaluation_result,
                item.has_ssl,
                item.is_always_ssl,
                item.max_depth if item.max_depth is not None else "",
                item.svcmd,
                item.site_structure,
                item.total_pages if item.total_pages is not None else "",
                item.cms_name,
                item.description,
                item.contact_fields,
                item.rejection_reason,
                item.remarks,
                item.operator_name,
            ]
            ws.append(row_data)

            # 追加したデータ行にデザインを適用（上揃え + 罫線 + フォント）
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20  # 各データ行の高さにも少し余裕を持たせる

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border

                # 中央揃えにする列と、左揃え（上揃え）にする列を出し分ける
                # 日付, 調査結果, SSL, 階層, ページ, 担当 などは中央揃えが美しい
                if col_idx in [1, 3, 4, 5, 6, 9, 15]:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # --- 【自動調整】各列の幅をコンテンツの最大長に合わせて調整する ---
        # --- 【自動調整】各列の幅をコンテンツの最大長に合わせて調整する ---
        for col in ws.columns:
            max_len = 0
            # col[0].column が None の場合は処理をスキップ（型安全性の確保）
            col_num = col[0].column
            if col_num is None:
                continue

            col_letter = get_column_letter(col_num)

            for cell in col:
                # セル値の文字数を簡易カウント（Noneはスキップ、改行がある場合は一番長い行を基準に）
                if cell.value is not None:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        # 全角文字を2文字として簡易計算
                        val_len = sum(2 if ord(c) > 127 else 1 for c in line)
                        if val_len > max_len:
                            max_len = val_len

            # 少し余白（パディング）を足して、最低幅も保証する
            ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

        wb.save(str(output_path))
